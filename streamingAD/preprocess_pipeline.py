#!/usr/bin/env python3
"""
preprocess_pipeline.py — End-to-end preprocessing pipeline.

Two paths:
  A) Aligned data path (existing 25 movies):
     step04_01 (from aligned json) → step04_02 (ffmpeg split) → step04_03 (face align)

  B) Excel script path (new movies):
     step04_01_from_excel (parse xlsx) → step04_02 (ffmpeg split) → step04_03 (face align)

Progress via callback:  callback(stage_label, percent_float, message)
"""

from __future__ import annotations

import json
import re
import shlex
import shutil
import subprocess
import sys
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
STEP04_DIR = PROJECT_ROOT / "Step04_RunTest"
if str(STEP04_DIR) not in sys.path:
    sys.path.insert(0, str(STEP04_DIR))

DEFAULT_ALIGNED_JSON = PROJECT_ROOT / "Step03_AlignSript" / "aligned_from_refined.json"
DEFAULT_CMDQA_CSV = PROJECT_ROOT / "Step01_GetData" / "CMDQA.csv"
DEFAULT_OUTPUT_DIR = STEP04_DIR / "step04_final_by_movie_new"
DEFAULT_CLIPS_DIR = STEP04_DIR / "ad_clips_final"
DEFAULT_FACE_DIR = STEP04_DIR / "step04_03_face_align" / "json"

EPS = 1e-6
MAX_SCENE_GAP_SEC = 30.0

ProgressFn = Callable[[str, float, str], None]


def _sanitize_name(name: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_ ." else "_" for ch in str(name or "")).strip() or "unknown_movie"


def _to_float(value: Any, default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return default


def _parse_hhmmss_mmm(time_text: str) -> Optional[float]:
    text = str(time_text or "").strip()
    if not text:
        return None
    parts = text.split(":")
    if len(parts) != 3:
        return None
    hh, mm, ss_ms = parts
    if "," in ss_ms:
        ss_text, ms_text = ss_ms.split(",", 1)
    elif "." in ss_ms:
        ss_text, ms_text = ss_ms.split(".", 1)
    else:
        ss_text, ms_text = ss_ms, "0"
    try:
        return int(hh) * 3600 + int(mm) * 60 + int(ss_text) + int((ms_text + "000")[:3]) / 1000.0
    except ValueError:
        return None


def _sec_to_srt_time(total_sec: float) -> str:
    if total_sec < 0:
        total_sec = 0.0
    millis = int(round(total_sec * 1000))
    hh = millis // 3_600_000
    millis %= 3_600_000
    mm = millis // 60_000
    millis %= 60_000
    ss = millis // 1000
    ms = millis % 1000
    return f"{hh:02d}:{mm:02d}:{ss:02d},{ms:03d}"


def _canonical(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def canonical_movie_title(title: str) -> str:
    return _canonical(title)


def sanitize_filename(name: str) -> str:
    s = str(name or "").strip()
    s = re.sub(r"[\\/:*?\"<>|]", "_", s)
    s = re.sub(r"\s+", "_", s)
    return s or "unknown"


def dedupe_keep_order(items: Any) -> List[str]:
    seen = set()
    out = []
    try:
        iterator = iter(items)
    except (TypeError, ValueError):
        iterator = iter([])
    for it in iterator:
        s = str(it or "").strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


class _ExcelRow:
    __slots__ = ("row_index", "start_time", "end_time", "start_sec", "end_sec",
                 "dialog", "align_dialog", "record_type", "scene_index",
                 "characters", "location", "description")
    def __init__(self, **kw):
        for k, v in kw.items():
            setattr(self, k, v)


def _find_result_sheet(xlsx_path: Path) -> Optional[str]:
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    for name in wb.sheetnames:
        nl = name.lower()
        if nl.startswith("result") and nl not in ("subtitle", "script", "subtitle_en"):
            wb.close()
            return name
    wb.close()
    return None


def _load_xlsx_rows(xlsx_path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, object]]:
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        wb.close()
        return []
    headers = [str(v or "").strip() for v in header_row]
    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for idx, val in enumerate(row_vals):
            if idx < len(headers):
                row[headers[idx]] = val
        rows.append(row)
    wb.close()
    return rows


def generate_segs_from_excel(
    movie_title: str,
    video_path: str,
    xlsx_path: str,
    progress: ProgressFn,
    output_dir: Optional[Path] = None,
    max_gap_sec: float = MAX_SCENE_GAP_SEC,
) -> Optional[Path]:

    output_dir = Path(output_dir or DEFAULT_OUTPUT_DIR)
    out_file = output_dir / f"{sanitize_filename(movie_title)}.json"
    xlsx_p = Path(xlsx_path)

    progress("step04_01", 0.0, f"Parsing Excel: {xlsx_p.name}")

    result_sheet = _find_result_sheet(xlsx_p)
    if result_sheet:
        progress("step04_01", 0.0, f"Parsing Excel: {xlsx_p.name} [{result_sheet}]")
    raw_rows = _load_xlsx_rows(xlsx_p, sheet_name=result_sheet)
    if not raw_rows:
        progress("step04_01", 1.0, "Excel is empty")
        return None

    excel_rows: List[_ExcelRow] = []
    for i, row in enumerate(raw_rows):
        start_sec = _parse_hhmmss_mmm(str(row.get("start_time", "") or ""))
        end_sec = _parse_hhmmss_mmm(str(row.get("end_time", "") or ""))
        if start_sec is None or end_sec is None or end_sec < start_sec:
            continue
        excel_rows.append(_ExcelRow(
            row_index=i,
            start_time=str(row.get("start_time", "") or ""),
            end_time=str(row.get("end_time", "") or ""),
            start_sec=start_sec,
            end_sec=end_sec,
            dialog=str(row.get("dialog", "") or ""),
            align_dialog=str(row.get("align_dialog", "") or ""),
            record_type=str(row.get("record_type", "") or "dialog"),
            scene_index=str(row.get("scene_index", "") or "0"),
            characters=str(row.get("characters", "") or ""),
            location=str(row.get("location", "") or ""),
            description=str(row.get("description", "") or ""),
        ))

    if not excel_rows:
        progress("step04_01", 1.0, "No valid timed rows in Excel")
        return None

    progress("step04_01", 0.1, f"Loaded {len(excel_rows)} timed rows")

    # Group by scene_index
    scenes: Dict[str, List[_ExcelRow]] = {}
    for er in excel_rows:
        scenes.setdefault(er.scene_index, []).append(er)
    for rows in scenes.values():
        rows.sort(key=lambda r: r.start_sec)

    progress("step04_01", 0.2, f"Grouped into {len(scenes)} scenes")

    # Create AD segments: one per dialogue cluster within each scene
    ad_segments: List[Dict[str, object]] = []
    ad_idx = 0

    for scene_idx in sorted(scenes.keys(), key=lambda k: int(k) if k.isdigit() else 10**9):
        scene_rows = scenes[scene_idx]
        clusters: List[List[_ExcelRow]] = []
        current_cluster: List[_ExcelRow] = []

        for er in scene_rows:
            if not current_cluster:
                current_cluster.append(er)
            elif er.start_sec - current_cluster[-1].end_sec <= max_gap_sec:
                current_cluster.append(er)
            else:
                clusters.append(current_cluster)
                current_cluster = [er]
        if current_cluster:
            clusters.append(current_cluster)

        for cluster in clusters:
            ad_start = cluster[0].start_sec
            ad_end = cluster[-1].end_sec
            # Add padding
            ad_start = max(0.0, ad_start - 0.5)
            ad_end = ad_end + 0.5

            duration = ad_end - ad_start
            if duration < 0.5:
                continue

            ad_idx += 1

            matched_in_range: List[Dict[str, object]] = []
            for er in scene_rows:
                if er.start_sec >= ad_start - EPS and er.end_sec <= ad_end + EPS:
                    matched_in_range.append({
                        "Unnamed: 0": er.row_index,
                        "start_time": er.start_time,
                        "end_time": er.end_time,
                        "start_time_sec": er.start_sec,
                        "end_time_sec": er.end_sec,
                        "dialog": er.dialog,
                        "align_dialog": er.align_dialog,
                        "index_result": scene_idx,
                        "record_type": er.record_type,
                        "scene_index": er.scene_index,
                        "characters": er.characters,
                        "location": er.location,
                        "description": er.description,
                        "consistent": "1",
                    })

            aggregated = {
                "dialogs": dedupe_keep_order(r.dialog for r in cluster),
                "align_dialogs": dedupe_keep_order(r.align_dialog for r in cluster),
                "characters": dedupe_keep_order(r.characters for r in cluster),
                "locations": dedupe_keep_order(r.location for r in cluster),
                "descriptions": dedupe_keep_order(r.description for r in cluster),
                "record_types": dedupe_keep_order(r.record_type for r in cluster),
                "scene_indices": dedupe_keep_order(r.scene_index for r in cluster),
            }

            ad_id = f"{sanitize_filename(movie_title)}__clip{ad_idx:04d}__ad{ad_idx:04d}"

            seg = {
                "ad_id": ad_id,
                "movie_title": movie_title,
                "clip_index": ad_idx,
                "clip_file": video_path,
                "movie_file": video_path,
                "refined_time_sec": 0.0,
                "cmdqa": {
                    "row_id": 0, "movie_title": movie_title, "clip_index": 1,
                    "text": "", "scaled_start": 0.0, "scaled_end": duration,
                    "audiovault_start": 0.0, "audiovault_end": duration,
                    "duration": duration, "imdbid": "", "split": "custom",
                },
                "ad_movie_start_sec": ad_start,
                "ad_movie_end_sec": ad_end,
                "ad_movie_start_time": _sec_to_srt_time(ad_start),
                "ad_movie_end_time": _sec_to_srt_time(ad_end),
                "matched_rows_in_range_count": len(matched_in_range),
                "matched_rows_selected_count": len(matched_in_range),
                "matched_rows_selected": matched_in_range,
                "aggregated": aggregated,
            }
            ad_segments.append(seg)

    summary = {
        "movie_title": movie_title,
        "clip_count": ad_idx,
        "ad_segment_count": len(ad_segments),
        "warning_count": 0,
        "source": "excel_script",
    }

    payload = {
        "meta": {"generated_at": "", "source": "streamingAD", "mode": "excel"},
        "movie_summary": summary,
        "warnings": [],
        "ad_segments": ad_segments,
    }

    out_file.parent.mkdir(parents=True, exist_ok=True)
    with out_file.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    progress("step04_01", 1.0, f"Generated {len(ad_segments)} AD segments from {len(scenes)} scenes → {out_file}")
    return out_file


def load_aligned_items(aligned_json: Path) -> List[Dict[str, Any]]:
    import step04_01_final_data_organize as _s01
    return _s01.load_aligned_items(aligned_json)


def load_cmdqa_index(cmdqa_csv: Path) -> Dict:
    import step04_01_final_data_organize as _s01
    return _s01.load_cmdqa_index(cmdqa_csv)


def process_one_movie(
    movie_title: str, clip_items: List[Dict[str, Any]],
    cmdqa_index: Dict, min_context_rows: int,
) -> Dict[str, Any]:
    import step04_01_final_data_organize as _s01
    return _s01.process_one_movie(
        movie_title=movie_title, clip_items=clip_items,
        cmdqa_index=cmdqa_index, min_context_rows=min_context_rows,
    )


def write_json(out_file: Path, payload: Dict[str, Any]) -> None:
    import step04_01_final_data_organize as _s01
    _s01.write_json(out_file, payload)


def run_step04_01(
    movie_title: str,
    progress: ProgressFn,
    output_dir: Optional[Path] = None,
    aligned_json: Optional[Path] = None,
    cmdqa_csv: Optional[Path] = None,
    video_path: str = "",
    xlsx_path: str = "",
) -> Optional[Path]:

    output_dir = Path(output_dir or DEFAULT_OUTPUT_DIR)
    out_file = output_dir / f"{sanitize_filename(movie_title)}.json"

    if out_file.exists():
        progress("step04_01", 1.0, f"segment JSON already exists: {out_file}")
        return out_file

    # Path A: try aligned data first
    aligned_json_p = Path(aligned_json or DEFAULT_ALIGNED_JSON)
    if aligned_json_p.is_file():
        cmdqa_csv_p = Path(cmdqa_csv or DEFAULT_CMDQA_CSV)
        aligned_items = load_aligned_items(aligned_json_p)
        items_by_movie: Dict[str, List[Dict[str, Any]]] = {}
        for item in aligned_items:
            if not isinstance(item, dict):
                continue
            mt = (item.get("movie_title") or "").strip()
            if not mt:
                continue
            items_by_movie.setdefault(mt, []).append(item)

        matches = [(mt, rows) for mt, rows in items_by_movie.items()
                    if canonical_movie_title(mt) == canonical_movie_title(movie_title)]
        if matches:
            matched_title, rows = matches[0]
            progress("step04_01", 0.3, f"Using aligned data for '{matched_title}' ({len(rows)} clips)")
            cmdqa_index = load_cmdqa_index(cmdqa_csv_p)
            result = process_one_movie(
                movie_title=matched_title, clip_items=rows,
                cmdqa_index=cmdqa_index, min_context_rows=10,
            )
            payload = {
                "meta": {"generated_at": "", "source": "streamingAD"},
                "movie_summary": result["movie_summary"],
                "warnings": result["warnings"],
                "ad_segments": result["ad_segments"],
            }
            write_json(out_file, payload)
            seg_count = result["movie_summary"]["ad_segment_count"]
            progress("step04_01", 1.0, f"Generated {seg_count} AD segments → {out_file}")
            return out_file

    # Path B: generate from Excel script
    xp = (xlsx_path or "").strip()
    vp = (video_path or "").strip()
    if xp and vp:
        progress("step04_01", 0.0, f"No aligned data. Generating from Excel: {xp}")
        return generate_segs_from_excel(
            movie_title=movie_title, video_path=vp, xlsx_path=xp,
            progress=progress, output_dir=output_dir,
        )

    progress("step04_01", 1.0, f"No aligned data and no Excel script provided for '{movie_title}'")
    return None


def _resolve_input_clip(raw_path: str, base_dir: Path) -> Optional[Path]:
    text = str(raw_path or "").strip()
    if not text:
        return None
    p = Path(text)
    candidates = [p]
    if not p.is_absolute():
        candidates.append((base_dir / p).resolve())
        candidates.append((Path.cwd() / p).resolve())
    for cand in candidates:
        if cand.exists() and cand.is_file():
            return cand
    return None


def _run_cmd(cmd: Sequence[str]) -> Tuple[bool, str]:
    try:
        cp = subprocess.run(list(cmd), check=True, capture_output=True, text=True)
        return True, (cp.stderr or "").strip() or "ok"
    except subprocess.CalledProcessError as exc:
        return False, (exc.stderr or "").strip() or str(exc)


def run_step04_02(
    movie_title: str,
    video_path: str,
    seg_json: Path,
    progress: ProgressFn,
    clips_dir: Optional[Path] = None,
    ffmpeg_bin: str = "ffmpeg",
    force: bool = False,
) -> Optional[Path]:
    clips_dir = Path(clips_dir or DEFAULT_CLIPS_DIR)
    out_movie_dir = clips_dir / _sanitize_name(movie_title)

    if not seg_json.is_file():
        progress("step04_02", 1.0, f"segment JSON not found: {seg_json}")
        return None

    with seg_json.open("r", encoding="utf-8") as f:
        data = json.load(f)
    segments = data.get("ad_segments", []) if isinstance(data, dict) else []
    if not segments:
        progress("step04_02", 1.0, "No AD segments to split.")
        return out_movie_dir

    existing_count = 0
    if out_movie_dir.is_dir() and not force:
        existing_count = len(list(out_movie_dir.glob("*.mp4")))
    if existing_count >= len(segments):
        progress("step04_02", 1.0, f"All {existing_count} clips already exist in {out_movie_dir}")
        return out_movie_dir

    if shutil.which(ffmpeg_bin) is None and not Path(ffmpeg_bin).exists():
        progress("step04_02", 1.0, f"ffmpeg not found: {ffmpeg_bin}")
        return None

    total = len(segments)

    # Build task list: only segments needing extraction
    tasks = []
    skipped = 0
    vp = (video_path or "").strip()

    for idx, seg in enumerate(segments):
        if not isinstance(seg, dict):
            continue

        start_sec = _to_float(seg.get("ad_movie_start_sec"), default=0.0) or 0.0
        end_sec = _to_float(seg.get("ad_movie_end_sec"), default=0.0) or 0.0
        ad_id = str(seg.get("ad_id", "")).strip()
        clip_index = seg.get("clip_index")
        m = re.search(r"__clip(\d+)__ad(\d+)$", ad_id)
        if m:
            clip_num = int(m.group(1))
            ad_num = int(m.group(2))
        else:
            clip_num = int(_to_float(clip_index, default=0) or 0)
            ad_num = idx + 1

        out_path = out_movie_dir / f"clip{clip_num:04d}_ad{ad_num:04d}.mp4"

        if out_path.exists() and not force:
            skipped += 1
            continue

        duration = end_sec - start_sec
        if duration <= EPS:
            skipped += 1
            continue

        in_clip = None
        if vp:
            in_clip = _resolve_input_clip(vp, Path(vp).parent)
        if in_clip is None:
            clip_file_raw = str(seg.get("clip_file", "") or seg.get("movie_file", ""))
            in_clip = _resolve_input_clip(clip_file_raw, seg_json.resolve().parent)
        if in_clip is None:
            skipped += 1
            continue

        out_path.parent.mkdir(parents=True, exist_ok=True)
        cmd = [
            ffmpeg_bin, "-hide_banner", "-loglevel", "error", "-y",
            "-ss", f"{start_sec:.3f}",
            "-i", str(in_clip),
            "-t", f"{duration:.3f}",
            "-map", "0:v:0?", "-map", "0:a:0?",
            "-c:v", "mpeg4", "-q:v", "3",
            "-c:a", "aac", "-b:a", "128k",
            str(out_path),
        ]
        tasks.append((idx, cmd, out_path))

    pending = len(tasks)
    if pending == 0:
        progress("step04_02", 1.0, f"All {skipped} clips already exist in {out_movie_dir}")
        return out_movie_dir

    progress("step04_02", 0.0, f"Extracting {pending} clips with {max(1, min(8, pending))} parallel workers ...")

    ok = [0]
    failed = [0]
    done = [0]
    lock = threading.Lock()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    def _extract_one(args):
        _idx, _cmd, _out = args
        _success, _ = _run_cmd(_cmd)
        with lock:
            done[0] += 1
            if _success:
                ok[0] += 1
            else:
                failed[0] += 1
            pct = done[0] / pending
            progress("step04_02", pct, f"Clip {done[0]}/{pending}: {_out.name}")
        return _success

    max_workers = max(1, min(8, pending))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for _ in as_completed([ex.submit(_extract_one, t) for t in tasks]):
            pass

    ok_val, failed_val = ok[0], failed[0]
    progress("step04_02", 1.0, f"Done: {ok_val} ok, {skipped} skipped, {failed_val} failed → {out_movie_dir}")
    return out_movie_dir if ok_val > 0 or skipped > 0 else None


def run_step04_03(
    movie_title: str,
    progress: ProgressFn,
    gpu_id: int = 0,
    face_dir: Optional[Path] = None,
    clips_dir: Optional[Path] = None,
) -> Optional[Path]:
    face_dir = Path(face_dir or DEFAULT_FACE_DIR)
    clips_dir = Path(clips_dir or DEFAULT_CLIPS_DIR)
    step03_script = STEP04_DIR / "step04_03.py"

    out_movie_dir = face_dir / _sanitize_name(movie_title)
    if out_movie_dir.is_dir() and any(out_movie_dir.iterdir()):
        progress("step04_03", 1.0, f"face data already exists: {out_movie_dir}")
        return out_movie_dir

    if not step03_script.is_file():
        progress("step04_03", 1.0, f"step04_03 script not found at {step03_script}")
        return None

    clip_movie_dir = clips_dir / _sanitize_name(movie_title)
    if not clip_movie_dir.is_dir():
        progress("step04_03", 1.0, f"clips dir not found: {clip_movie_dir}")
        return None

    faces_db = STEP04_DIR / "step04_03_face_align" / "faces"
    plotree_db = STEP04_DIR / "faces" / "plotree"
    if not faces_db.is_dir() and not plotree_db.is_dir():
        progress("step04_03", 1.0, f"face gallery not found: {faces_db} or {plotree_db}")
        return None

    venv_python = "/mnt/disk6new/wzq/env/videollava/bin/python"

    cmd = [
        venv_python, str(step03_script),
        "--ad-clips-dir", str(clips_dir),
        "--faces-movies-dir", str(faces_db),
        "--output-dir", str(face_dir.parent),
        "--gpu-id", str(gpu_id),
        "--movie", movie_title,
    ]

    progress("step04_03", 0.0, f"Starting face alignment for '{movie_title}'...")
    process = subprocess.Popen(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, bufsize=1,
        env={**__import__("os").environ,
             "HF_HOME": "/tmp/hf_cache",
             "TRANSFORMERS_OFFLINE": "1",
             "HF_DATASETS_OFFLINE": "1"},
    )
    assert process.stdout
    last_pct = 0
    for line in process.stdout:
        line = line.rstrip()
        m = re.search(r"(\d+)%|(\d+)/(\d+)\s", line)
        if m:
            if m.group(1):
                pct = int(m.group(1)) / 100
            else:
                pct = int(m.group(2)) / max(int(m.group(3)), 1)
            if pct > last_pct:
                last_pct = pct
                progress("step04_03", pct, line[:120])
    process.wait()
    if process.returncode == 0:
        progress("step04_03", 1.0, f"Face alignment complete → {out_movie_dir}")
        return out_movie_dir
    else:
        progress("step04_03", 1.0, f"Face alignment exited with code {process.returncode}")
        return None


def run_full_preprocess(
    movie_title: str,
    progress: ProgressFn,
    gpu_id: int = 0,
    force: bool = False,
    video_path: str = "",
    xlsx_path: str = "",
) -> Dict[str, Any]:

    result: Dict[str, Any] = {
        "movie_title": movie_title,
        "seg_json": None,
        "clips_dir": None,
        "face_dir": None,
        "status": "unknown",
        "messages": [],
    }

    seg_json = run_step04_01(
        movie_title=movie_title, progress=progress,
        output_dir=DEFAULT_OUTPUT_DIR,
        video_path=video_path, xlsx_path=xlsx_path,
    )
    if seg_json is None:
        result["status"] = "no_segment_data"
        result["messages"].append(
            f"Failed to generate segment data. Provide an Excel script path for new movies."
        )
        return result
    result["seg_json"] = str(seg_json)

    clips_dir = run_step04_02(
        movie_title=movie_title, video_path=video_path,
        seg_json=seg_json, progress=progress,
        clips_dir=DEFAULT_CLIPS_DIR, force=force,
    )
    if clips_dir is None:
        result["status"] = "no_clips"
        result["messages"].append("Failed to split clips. Check ffmpeg and video path.")
        return result
    result["clips_dir"] = str(clips_dir)

    face_dir = run_step04_03(
        movie_title=movie_title, progress=progress,
        gpu_id=gpu_id, face_dir=DEFAULT_FACE_DIR, clips_dir=DEFAULT_CLIPS_DIR,
    )
    result["face_dir"] = str(face_dir) if face_dir else None
    if face_dir is None:
        result["messages"].append("Face alignment skipped or failed (non-blocking)")

    result["status"] = "ready"
    result["messages"].append(f"All preprocessing complete for '{movie_title}'")
    return result
