#!/usr/bin/env python3
"""
pipeline_script_based.py — Method 1: Script-Based Audio Description Generation
================================================================================

Uses time-aligned script (xlsx) to:
  1. Parse dialogue rows with character names, locations, scene indices
  2. Detect dialogue gaps from script timing (end of one line → start of next)
  3. Clip video segments for each gap
  4. Build rich context (characters, location, surrounding dialogue)
  5. Generate AD text via Video-LLaMA

Requirements:
  - Video files in video_dir
  - Matching xlsx aligned scripts in xlsx_dir
  - Video-LLaMA model (loaded from models/)

Output: JSON per movie, same format as Method 2 for direct comparison.

Usage:
    conda run -n videollava python streamingAD/pipeline_script_based.py \
        --video-dir /mnt/disk1new/storyvideo/Movie \
        --xlsx-dir /mnt/disk1new/storyvideo/Alignedscript/Movie \
        --output-dir /mnt/disk1new/ylz/newAD/compare/script_based \
        --gap-threshold 4.0 --gpu-id 0 \
        --only-movie "Shawshank"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

PROJECT_ROOT = Path(__file__).resolve().parents[1]
STREAMING_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(STREAMING_ROOT))

# ── Default Config ───────────────────────────────────────────
VIDEO_DIRS: List[str] = ["/mnt/disk1new/storyvideo/Movie"]
XLSX_DIRS: List[str] = ["/mnt/disk1new/storyvideo/Alignedscript/Movie"]
GT_CSV: str = "/mnt/disk1new/ylz/newAD/AutoAD3/cmd_ad_anno_v1.csv"
DEFAULT_OUTPUT_DIR: str = "/mnt/disk1new/ylz/newAD/compare/script_based"
DEFAULT_GAP_SEC: float = 4.0
DEFAULT_TEMPERATURE: float = 0.2
DEFAULT_MAX_TOKENS: int = 256
DEFAULT_GPU: int = 0

TASK_PROMPT = (
    "Describe what is happening in this clip concisely. "
    "Focus on visible actions, movements, and expressions. "
    "If character names are mentioned in the context, use them "
    "(e.g. 'Don Vito Corleone walks...' not 'A man walks...'). "
    "Do not quote dialogue."
)
# ──────────────────────────────────────────────────────────────


def _pct(a, b): return f"{a}/{b}" if b else "0"
def _fmt_sec(s): return f"{s:.1f}s"


@dataclass
class DialogueRow:
    row_index: int
    start_sec: float
    end_sec: float
    start_time: str
    end_time: str
    dialog: str
    scene_index: str
    characters: str
    location: str


@dataclass
class GapCandidate:
    gap_id: int
    scene_index: str
    location: str
    gap_start_sec: float
    gap_end_sec: float
    gap_start_time: str
    gap_end_time: str
    gap_duration_sec: float
    context_before: List[str]
    context_after: List[str]
    characters: List[str]


# ── Helpers ──────────────────────────────────────────────────

def _parse_hhmmss_mmm(time_text: str) -> Optional[float]:
    text = str(time_text or "").strip()
    if not text: return None
    parts = text.split(":")
    if len(parts) != 3: return None
    hh, mm, ss_ms = parts
    if "," in ss_ms:
        s, ms = ss_ms.split(",", 1)
    elif "." in ss_ms:
        s, ms = ss_ms.split(".", 1)
    else:
        s, ms = ss_ms, "0"
    try:
        return int(hh) * 3600 + int(mm) * 60 + int(s) + int((ms + "000")[:3]) / 1000.0
    except ValueError:
        return None


def _sec_to_timestamp(seconds: float) -> str:
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"


def _clean_name(fname: str) -> str:
    name = os.path.splitext(fname)[0]
    return re.sub(r'^(IMDB-\d+-|douban-\d+-)', '', name).strip().lower()


def _find_result_sheet(xlsx_path: Path) -> Optional[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn.lower() in ("result", "results", "aligned"):
                return sn
        return wb.sheetnames[0] if wb.sheetnames else None
    except Exception:
        return None


def _load_xlsx_rows(xlsx_path: Path, sheet_name: Optional[str] = None):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(rows_iter, [])]
    result = []
    for vals in rows_iter:
        row = {}
        for i, v in enumerate(vals):
            if i < len(header):
                row[header[i]] = v
        result.append(row)
    return result


# ── Matching ─────────────────────────────────────────────────

def match_videos_to_xlsxs(
    video_dirs: List[str],
    xlsx_dirs: List[str],
    gt_csv: Optional[str] = None,
    only_movie: Optional[str] = None,
) -> List[Tuple[Path, Path]]:
    import csv as _csv
    
    video_names: Dict[str, Path] = {}
    for vd in video_dirs:
        vp = Path(vd)
        if vp.is_dir():
            for f in vp.iterdir():
                if f.suffix.lower() in {".mp4", ".mkv", ".avi", ".mov"}:
                    video_names[_clean_name(f.name)] = f

    xlsx_names: Dict[str, Path] = {}
    for xd in xlsx_dirs:
        xp = Path(xd)
        if xp.is_dir():
            for f in xp.iterdir():
                if f.suffix.lower() == ".xlsx":
                    xlsx_names[_clean_name(f.name)] = f

    gt_titles: set = set()
    if gt_csv and Path(gt_csv).exists():
        with open(gt_csv) as f:
            for row in _csv.DictReader(f):
                t = row.get('movie_title', '').strip().lower()
                if t: gt_titles.add(t)

    pairs: List[Tuple[Path, Path]] = []
    for vname, vp in sorted(video_names.items()):
        if vname not in xlsx_names: continue
        if gt_titles and vname not in gt_titles: continue
        pairs.append((vp.resolve(), xlsx_names[vname].resolve()))

    if only_movie:
        pairs = [(v, x) for v, x in pairs if only_movie.lower() in v.stem.lower()]

    return pairs


# ── Dialogue Parsing ─────────────────────────────────────────

def parse_dialogue_rows(xlsx_path: Path) -> List[DialogueRow]:
    result_sheet = _find_result_sheet(xlsx_path)
    raw_rows = _load_xlsx_rows(xlsx_path, sheet_name=result_sheet)
    rows: List[DialogueRow] = []
    for i, raw in enumerate(raw_rows):
        start_sec = _parse_hhmmss_mmm(str(raw.get("start_time", "") or ""))
        end_sec = _parse_hhmmss_mmm(str(raw.get("end_time", "") or ""))
        if start_sec is None or end_sec is None or end_sec < start_sec:
            continue
        dialog = str(raw.get("dialog", "") or "").strip()
        if not dialog: continue
        rows.append(DialogueRow(
            row_index=i,
            start_sec=start_sec,
            end_sec=end_sec,
            start_time=str(raw.get("start_time", "") or ""),
            end_time=str(raw.get("end_time", "") or ""),
            dialog=dialog,
            scene_index=str(raw.get("scene_index", "") or "0"),
            characters=str(raw.get("characters", "") or ""),
            location=str(raw.get("location", "") or ""),
        ))
    rows.sort(key=lambda r: r.start_sec)
    return rows


def detect_dialogue_gaps(
    rows: List[DialogueRow],
    gap_threshold_sec: float = DEFAULT_GAP_SEC,
    max_context_lines: int = 5,
) -> List[GapCandidate]:
    scenes: Dict[str, List[DialogueRow]] = {}
    for r in rows:
        scenes.setdefault(r.scene_index, []).append(r)
    for srows in scenes.values():
        srows.sort(key=lambda r: r.start_sec)

    sorted_keys = sorted(scenes.keys(), key=lambda k: int(k) if k.isdigit() else 10**9)
    candidates: List[GapCandidate] = []
    gap_id = 0

    def _add_gap(sa, ia, sb, ib, sc_a, sc_b):
        nonlocal gap_id
        gs = sa[ia].end_sec
        ge = sb[ib].start_sec
        dur = ge - gs
        if dur < gap_threshold_sec: return

        ctx_before = [r.dialog for r in sa[max(0, ia - max_context_lines + 1):ia + 1]]
        ctx_after = [r.dialog for r in sb[ib:min(len(sb), ib + max_context_lines)]]

        chars_set = set()
        for r in sa[max(0, ia - 2):min(len(sa), ia + 1)]:
            for c in r.characters.split(","):
                if c.strip(): chars_set.add(c.strip())
        for r in sb[ib:min(len(sb), ib + 3)]:
            for c in r.characters.split(","):
                if c.strip(): chars_set.add(c.strip())

        loc = sa[ia].location or sb[ib].location or ""
        sc_label = sc_a if sc_a == sc_b else f"{sc_a}→{sc_b}"

        gap_id += 1
        candidates.append(GapCandidate(
            gap_id=gap_id, scene_index=sc_label, location=loc,
            gap_start_sec=gs, gap_end_sec=ge,
            gap_start_time=_sec_to_timestamp(gs), gap_end_time=_sec_to_timestamp(ge),
            gap_duration_sec=dur, context_before=ctx_before, context_after=ctx_after,
            characters=sorted(chars_set),
        ))

    for sc_idx in sorted_keys:
        sr = scenes[sc_idx]
        for i in range(len(sr) - 1):
            _add_gap(sr, i, sr, i + 1, sc_idx, sc_idx)

    for si in range(len(sorted_keys) - 1):
        a, b = sorted_keys[si], sorted_keys[si + 1]
        _add_gap(scenes[a], len(scenes[a]) - 1, scenes[b], 0, a, b)

    return candidates


def build_gap_context(candidate: GapCandidate) -> str:
    parts: List[str] = ["[Scene context]"]
    if candidate.location: parts.append(f"Location: {candidate.location}")
    if candidate.characters: parts.append(f"Characters: {', '.join(candidate.characters)}")
    if candidate.context_before:
        parts.append("\n[Dialogue before the gap]")
        for d in candidate.context_before: parts.append(f"- {d}")
    if candidate.context_after:
        parts.append("\n[Dialogue after the gap]")
        for d in candidate.context_after: parts.append(f"- {d}")
    parts.append(f"\n[Gap description]\nThis is a silent interval ({candidate.gap_duration_sec:.1f}s) "
                  f"between two dialogue segments in scene {candidate.scene_index}.")
    return "\n".join(parts)


# ── Video Processing ─────────────────────────────────────────

def extract_clip(video_path: Path, start_sec: float, end_sec: float,
                 output_path: Path, ffmpeg_bin: str = "ffmpeg") -> bool:
    duration = end_sec - start_sec
    if duration <= 0: return False
    output_path.parent.mkdir(parents=True, exist_ok=True)
    cmd = [
        ffmpeg_bin, "-hide_banner", "-loglevel", "error", "-y",
        "-ss", f"{start_sec:.3f}", "-i", str(video_path),
        "-t", f"{duration:.3f}", "-map", "0:v:0?", "-map", "0:a:0?",
        "-c", "copy", "-avoid_negative_ts", "make_zero", str(output_path),
    ]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
        return True
    except subprocess.CalledProcessError:
        return False


def _prepare_fast_video(src_video: Path, cache_dir: Path,
                        ffmpeg_bin: str = "ffmpeg") -> Path:
    fast_video = cache_dir / f"fast_{src_video.stem}.mp4"
    if fast_video.exists(): return fast_video
    cmd = [
        ffmpeg_bin, "-hide_banner", "-loglevel", "error", "-y",
        "-i", str(src_video), "-map", "0:v:0?", "-map", "0:a:0?",
        "-c", "copy", "-movflags", "+faststart", str(fast_video),
    ]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
        return fast_video
    except subprocess.CalledProcessError as exc:
        stderr = (exc.stderr or "").strip()
        print(f"[warn] fast video prep failed for {src_video.name}; fallback to source video", flush=True)
        if stderr:
            print(f"[warn] ffmpeg: {stderr.splitlines()[-1]}", flush=True)
        try:
            if fast_video.exists():
                fast_video.unlink()
        except Exception:
            pass
        return src_video


# ── Main Pipeline ────────────────────────────────────────────

def process_one_movie(
    engine,
    video_path: Path,
    xlsx_path: Path,
    output_dir: Path,
    gap_threshold_sec: float,
    temperature: float,
    max_new_tokens: int,
) -> Dict[str, Any]:
    movie_name = video_path.stem
    print(f"\n{'='*60}")
    print(f"[Method 1: Script-Based] {movie_name}")
    print(f"  Video: {video_path}")
    print(f"  Script: {xlsx_path}")
    print(f"{'='*60}")

    t_start = time.monotonic()

    rows = parse_dialogue_rows(xlsx_path)
    if not rows:
        return {"movie": movie_name, "method": "script_based", "status": "no_dialogue", "ad_entries": []}

    print(f"  {len(rows)} dialogue rows, {len(set(r.scene_index for r in rows))} scenes")

    candidates = detect_dialogue_gaps(rows, gap_threshold_sec)
    if not candidates:
        return {"movie": movie_name, "method": "script_based", "status": "no_gaps",
                "gap_threshold_sec": gap_threshold_sec, "ad_entries": []}

    print(f"  {len(candidates)} candidate gaps (>= {gap_threshold_sec}s)")

    # ══════ Phase A4: Face gallery ══════
    from face_gallery import load_gallery, lookup_face_images
    gallery_embs, gallery_people = load_gallery(movie_name)
    if gallery_embs is not None:
        print(f"  [Phase A4] Face gallery: {len(gallery_people)} characters loaded")
    else:
        print(f"  [Phase A4] Face gallery: not available for {movie_name}")

    clip_tmp = output_dir / f".tmp_{movie_name}"
    clip_tmp.mkdir(parents=True, exist_ok=True)
    fast_video = _prepare_fast_video(video_path, clip_tmp)

    entries: List[Dict[str, Any]] = []
    inference_total = 0.0
    extract_total = 0.0

    face_avatar_cache: Dict[str, List[Path]] = {}

    for idx, cand in enumerate(candidates):
        clip_path = clip_tmp / f"gap{cand.gap_id:04d}.mp4"

        t0 = time.monotonic()
        if not extract_clip(fast_video, cand.gap_start_sec, cand.gap_end_sec, clip_path):
            print(f"  [{_pct(idx+1, len(candidates))}] Gap {cand.gap_id}: extract FAILED")
            continue
        extract_time = time.monotonic() - t0
        extract_total += extract_time

        context_text = build_gap_context(cand)

        face_avatars: List[Path] = []
        if gallery_people is not None and cand.characters:
            cache_key = "|".join(sorted(cand.characters))
            if cache_key not in face_avatar_cache:
                face_avatar_cache[cache_key] = lookup_face_images(
                    cand.characters, gallery_people, movie_name,
                )
            face_avatars = face_avatar_cache[cache_key]

        try:
            ad_text, inf_time, _ = engine.infer_one_segment(
                clip_path=clip_path, context_text=context_text,
                task_prompt=TASK_PROMPT, temperature=temperature, max_new_tokens=max_new_tokens,
                face_avatars=face_avatars, character_names=cand.characters,
            )
        except Exception as e:
            ad_text = f"[ERROR: {e}]"
            inf_time = 0.0
        inference_total += inf_time

        print(f"  [{_pct(idx+1, len(candidates))}] Gap {cand.gap_id} ({cand.gap_duration_sec:.1f}s): {ad_text}")

        entries.append({
            "gap_id": cand.gap_id, "scene_index": cand.scene_index,
            "location": cand.location, "characters": cand.characters,
            "gap_start_time": cand.gap_start_time, "gap_end_time": cand.gap_end_time,
            "gap_start_sec": round(cand.gap_start_sec, 3),
            "gap_end_sec": round(cand.gap_end_sec, 3),
            "gap_duration_sec": round(cand.gap_duration_sec, 1),
            "context_before": cand.context_before, "context_after": cand.context_after,
            "ad_text": ad_text,
            "inference_time_sec": round(inf_time, 1),
            "extract_time_sec": round(extract_time, 1),
        })

        if clip_path.exists():
            clip_path.unlink()

    shutil.rmtree(str(clip_tmp), ignore_errors=True)

    t_total = time.monotonic() - t_start

    result = {
        "movie": movie_name, "method": "script_based",
        "video_path": str(video_path), "xlsx_path": str(xlsx_path),
        "gap_threshold_sec": gap_threshold_sec,
        "total_gaps": len(candidates), "generated_count": len(entries),
        "preprocess_time_sec": round(t_total - inference_total, 1),
        "inference_total_time_sec": round(inference_total, 1),
        "extract_total_time_sec": round(extract_total, 1),
        "total_time_sec": round(t_total, 1),
        "ad_entries": entries,
    }

    out_file = output_dir / f"{movie_name}_ad_output.json"
    with out_file.open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"  ✅ Saved: {out_file}")
    print(f"  Summary: {len(entries)}/{len(candidates)} ADs in {_fmt_sec(t_total)}")
    return result


# ── Entry Point ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Method 1: Script-Based AD Generation")
    parser.add_argument("--video-dir", nargs="*", default=VIDEO_DIRS)
    parser.add_argument("--xlsx-dir", nargs="*", default=XLSX_DIRS)
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--gap-threshold", type=float, default=DEFAULT_GAP_SEC)
    parser.add_argument("--temperature", type=float, default=DEFAULT_TEMPERATURE)
    parser.add_argument("--max-new-tokens", type=int, default=DEFAULT_MAX_TOKENS)
    parser.add_argument("--gpu-id", type=int, default=DEFAULT_GPU)
    parser.add_argument("--only-movie", default=None)
    parser.add_argument("--no-gt-filter", action="store_true")
    args = parser.parse_args()

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    gt_csv = None if args.no_gt_filter else GT_CSV
    pairs = match_videos_to_xlsxs(args.video_dir, args.xlsx_dir,
                                   gt_csv=gt_csv, only_movie=args.only_movie)
    if not pairs:
        print("ERROR: No video-xlsx pairs found."); sys.exit(1)

    print("=" * 60)
    print("Method 1: Script-Based AD Generation")
    print(f"  Pairs:     {len(pairs)}")
    print(f"  Gap min:   {args.gap_threshold}s")
    print(f"  Output:    {output_dir}")
    print(f"  GPU:       {args.gpu_id}")
    print("=" * 60)

    print("\n--- Loading Video-LLaMA ---")
    from ad_engine import build_ad_engine
    engine = build_ad_engine(gpu_id=args.gpu_id)
    print()

    all_results = []
    for vi, (vp, xp) in enumerate(pairs):
        print(f"\n[{_pct(vi+1, len(pairs))}] {vp.stem}")
        try:
            r = process_one_movie(
                engine, video_path=vp, xlsx_path=xp, output_dir=output_dir,
                gap_threshold_sec=args.gap_threshold, temperature=args.temperature,
                max_new_tokens=args.max_new_tokens,
            )
            all_results.append(r)
            print(f"  [{_pct(vi+1, len(pairs))}] {r['movie']}: "
                  f"{r['generated_count']}/{r['total_gaps']} ADs")
        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()

    total_g = sum(r.get("total_gaps", 0) for r in all_results)
    total_a = sum(r.get("generated_count", 0) for r in all_results)
    print(f"\n=== Done: {total_a}/{total_g} ADs across {len(all_results)} movies ===")
    print(f"Output: {output_dir}")


if __name__ == "__main__":
    main()
